"""
Elaw Carrefour — Indicação de Prepostos em Lote

Notas técnicas (JSF/PrimeFaces):
- Navegação direta por URL não funciona — usar busca global
- Botões via JavaScript (podem estar fora do viewport)
- Autocompletes PrimeFaces: foco via JS + digitação real (keyboard.type)
- page.evaluate com string NÃO aceita return no top-level — usar IIFEs
"""

import os
import re
import time
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Garante que build e runtime usam o mesmo caminho para o Chromium no Render.
# Deve ser definido ANTES de importar o playwright.
if os.environ.get("RENDER"):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "/opt/render/project/src/.playwright-browsers"

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

ELAW_URL      = "https://carrefour.elaw.com.br"
LOGIN_TIMEOUT = 120_000
PAGE_TIMEOUT  = 40_000   # Render é lento — 40s para carregamentos de página
POLL_ATTEMPTS = 12        # 12 × 2s = 24s máx de polling para autocompletes
POLL_WAIT     = 2.0

IS_SERVER = bool(os.environ.get("RENDER") or os.environ.get("IS_SERVER"))


# ── Entry point ───────────────────────────────────────────────────────────────

def run_automation(rows: list[dict], log, report_path: Path, state: dict | None = None):
    results = []
    novos   = 0

    with sync_playwright() as p:
        if IS_SERVER:
            browser = _launch_server(p)
            ctx  = browser.new_context(viewport={"width": 1280, "height": 900})
            page = ctx.new_page()
        else:
            chrome_profile = str(Path(__file__).parent / "chrome_profile")
            ctx  = p.chromium.launch_persistent_context(
                chrome_profile,
                headless=False,
                viewport={"width": 1280, "height": 900},
                slow_mo=120,
            )
            page = ctx.pages[0] if ctx.pages else ctx.new_page()

        log("Abrindo Elaw Carrefour...")
        page.goto(ELAW_URL, wait_until="networkidle", timeout=30_000)

        if _is_login_page(page):
            if IS_SERVER:
                log("Fazendo login automático...", "info")
                _auto_login(page, log)
            else:
                log("⚠️ Sessão expirada — faça login no browser aberto.", "warn")
                page.wait_for_url(f"**{ELAW_URL}/**", timeout=LOGIN_TIMEOUT)
                page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
                log("✅ Login detectado, iniciando automação...")
        else:
            log("✅ Sessão ativa, iniciando automação...")

        # Navega para processoList onde a barra de busca está sempre disponível
        page.goto(f"{ELAW_URL}/processoList.elaw", wait_until="networkidle", timeout=PAGE_TIMEOUT)
        page.wait_for_selector('[id*="globaSearchAutocomplete_input"]', state="visible", timeout=PAGE_TIMEOUT)

        total = len(rows)
        for i, row in enumerate(rows, 1):
            # ── Pause ──────────────────────────────────────────────────────────
            if state and state.get("paused"):
                log("⏸ Pausado — aguardando retomada...", "warn")
                while state.get("paused"):
                    time.sleep(1)
                log("▶️ Retomando...", "info")

            numero    = str(row.get("numero_processo", "")).strip()
            nome      = str(row.get("nome_preposto", "")).strip()
            cpf       = re.sub(r"\D", "", str(row.get("cpf_preposto") or ""))
            email_raw = str(row.get("email_preposto") or "").strip()
            email     = email_raw.replace(";", " ").split()[0] if email_raw else ""
            telefone  = str(row.get("telefone_preposto") or "").strip()

            log(f"[{i}/{total}] {numero} — {nome}...")

            for attempt in range(2):
                try:
                    status, detail, is_novo = _process_row(
                        page, numero, nome, cpf, email, telefone
                    )
                    if is_novo:
                        novos += 1
                    break
                except Exception as e:
                    err_str = str(e)
                    if "Execution context was destroyed" in err_str and attempt == 0:
                        log(f"  🔁 Contexto destruído, tentando novamente...", "warn")
                        _recover_page(page, log)
                        # continua para a segunda tentativa
                    else:
                        status, detail, is_novo = "ERRO", err_str[:300], False
                        _recover_page(page, log)
                        break

            row_result = {
                "numero_processo": numero,
                "nome_preposto":   nome,
                "cpf":             cpf,
                "email":           email,
                "status":          status,
                "detalhe":         detail,
                "horario":         datetime.now().strftime("%H:%M:%S"),
            }
            results.append(row_result)
            if state is not None:
                state["results"].append(row_result)

            icons = {"OK": "✅", "JÁ CONFIRMADO": "ℹ️", "ERRO": "❌"}
            css   = {"OK": "ok", "JÁ CONFIRMADO": "ok", "ERRO": "error"}
            log(f"  {icons.get(status,'❌')} {status}: {detail}", css.get(status, "error"))

        if IS_SERVER:
            browser.close()
        else:
            ctx.close()

    _build_report(results, report_path)
    ok = sum(1 for r in results if r["status"] in ("OK", "JÁ CONFIRMADO"))
    log(
        f"Concluído: {ok}/{len(results)} processos com sucesso "
        f"({novos} preposto(s) novo(s) cadastrado(s)).",
        "done",
    )


# ── Browser helpers ───────────────────────────────────────────────────────────

def _launch_server(p):
    return p.chromium.launch(
        headless=True,
        args=[
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
            "--disable-setuid-sandbox",
            "--single-process",
        ],
    )


def _auto_login(page, log):
    elaw_user = os.environ.get("ELAW_USER", "")
    elaw_pass = os.environ.get("ELAW_PASS", "")

    if not elaw_user or not elaw_pass:
        raise Exception(
            "Variáveis ELAW_USER e ELAW_PASS não configuradas. "
            "Adicione-as nas variáveis de ambiente do Render."
        )

    # Espera os campos estarem prontos (headless pode renderizar após domcontentloaded)
    page.wait_for_selector("#username", state="visible", timeout=PAGE_TIMEOUT)
    page.wait_for_selector("#authKey",  state="visible", timeout=PAGE_TIMEOUT)

    log("Preenchendo credenciais Elaw...")
    page.fill("#username", elaw_user, timeout=PAGE_TIMEOUT)
    page.fill("#authKey",  elaw_pass, timeout=PAGE_TIMEOUT)

    # Tentar clicar com force=True (ignora checagens de visibilidade do Playwright)
    try:
        page.locator("button.ui-button").first.click(force=True, timeout=PAGE_TIMEOUT)
    except Exception:
        # Fallback: submeter via JS
        page.evaluate("""(() => {
            const btn = document.querySelector('button.ui-button') ||
                        Array.from(document.querySelectorAll('button'))
                            .find(b => b.textContent.trim().includes('Acessar'));
            if (btn) btn.click();
            else { const f = document.querySelector('form'); if (f) f.submit(); }
        })()""")

    page.wait_for_load_state("networkidle", timeout=30_000)

    if _is_login_page(page):
        page.screenshot(path="/tmp/debug_login.png", full_page=True)
        err_msg = page.evaluate("""(() => {
            const el = document.querySelector(
                '.ui-messages-error-summary, .ui-messages-error, [class*="error-msg"], .growl-message'
            );
            return el ? el.textContent.trim() : null;
        })()""")
        detail = f": {err_msg}" if err_msg else " — verifique ELAW_USER e ELAW_PASS no Render."
        raise Exception(f"Login falhou{detail} (screenshot salvo em /debug-screenshot)")

    log("✅ Login concluído.")


def _recover_page(page, log):
    try:
        page.goto(f"{ELAW_URL}/processoList.elaw", wait_until="networkidle", timeout=PAGE_TIMEOUT)
        page.wait_for_selector(
            '[id*="globaSearchAutocomplete_input"]',
            state="visible",
            timeout=PAGE_TIMEOUT,
        )
    except Exception as recover_err:
        log(f"  ⚠️ Recovery falhou: {str(recover_err)[:120]}", "warn")


def _is_login_page(page) -> bool:
    url = page.url.lower()
    if any(k in url for k in ("login", "signin", "sso", "auth", "microsoftonline")):
        return True
    try:
        return bool(page.evaluate("""(() => {
            const isVisible = el => el && el.offsetParent !== null
                && getComputedStyle(el).display !== 'none'
                && getComputedStyle(el).visibility !== 'hidden';
            return isVisible(document.getElementById('username'))
                || isVisible(document.getElementById('authKey'));
        })()"""))
    except Exception:
        return False


# ── Fluxo por processo ────────────────────────────────────────────────────────

def _process_row(page, numero, nome, cpf, email, telefone):
    _navigate_to_process(page, numero)

    task_status = _click_task_confirm(page)
    if task_status == "ja_confirmado":
        return "JÁ CONFIRMADO", "Tarefa já estava concluída anteriormente", False

    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)
    is_novo = _fill_preposto(page, nome, cpf, email, telefone)

    page.evaluate("document.getElementById('btnConfirmaSim').click()")
    time.sleep(2)
    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)

    detail = (
        "Preposto indicado com sucesso (novo cadastro)"
        if is_novo else
        "Preposto indicado com sucesso"
    )
    return "OK", detail, is_novo


# ── Navegação ─────────────────────────────────────────────────────────────────

def _navigate_to_process(page, numero):
    page.wait_for_selector('[id*="globaSearchAutocomplete_input"]', timeout=PAGE_TIMEOUT)

    for attempt in range(2):
        page.evaluate("""
            const el = document.querySelector('[id*="globaSearchAutocomplete_input"]');
            el.value = '';
            el.focus();
            el.click();
        """)
        time.sleep(0.4)
        page.keyboard.type(numero, delay=65)

        clicked = False
        for _ in range(POLL_ATTEMPTS):
            time.sleep(POLL_WAIT)
            result = page.evaluate("""(() => {
                const panel = document.querySelector('[id$="globaSearchAutocomplete_panel"]');
                const items = panel ? panel.querySelectorAll('li') : [];
                if (items.length > 0) { items[0].click(); return 'clicado'; }
                return 'vazio';
            })()""")
            if result == "clicado":
                clicked = True
                break

        if clicked:
            break

    if not clicked:
        raise Exception(f"Autocomplete da busca não abriu para o processo {numero}")

    try:
        page.wait_for_url("**/processoView.elaw**", timeout=PAGE_TIMEOUT)
    except PWTimeout:
        raise Exception(f"Processo {numero} não encontrado no sistema")

    page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT)


# ── Tarefa ────────────────────────────────────────────────────────────────────

def _click_task_confirm(page):
    result = page.evaluate("""(() => {
        for (const row of document.querySelectorAll('tr')) {
            if (row.textContent.includes('Indicar Preposto')) {
                const btn = row.querySelector('button[id*="confirmAgendamento"]');
                if (btn) { btn.click(); return btn.id; }
                return 'btn_nao_encontrado';
            }
        }
        return 'tarefa_nao_encontrada';
    })()""")

    if result == "tarefa_nao_encontrada":
        return "ja_confirmado"

    if result == "btn_nao_encontrado":
        raise Exception("Botão confirmAgendamento não encontrado na linha da tarefa")

    try:
        page.wait_for_url("**/agendamentoContenciosoConfirm.elaw**", timeout=PAGE_TIMEOUT)
    except PWTimeout:
        # Algumas versões do Elaw carregam o form sem mudar a URL — verificar pelo campo
        try:
            page.wait_for_selector('[id$="pgAutoPreposto_input"]', state="visible", timeout=PAGE_TIMEOUT)
        except PWTimeout:
            # Salva screenshot + URL atual para diagnóstico
            try:
                page.screenshot(path="/tmp/debug_confirm.png", full_page=True)
            except Exception:
                pass
            current_url = page.url
            raise Exception(
                f"Tela de confirmação de agendamento não carregou "
                f"(URL atual: {current_url}) — screenshot em /debug-confirm-screenshot"
            )

    return "ok"


# ── Preposto ──────────────────────────────────────────────────────────────────

def _fill_preposto(page, nome, cpf, email, telefone) -> bool:
    search_term = " ".join(nome.split()[:3])
    items = _type_and_poll_autocomplete(page, search_term)

    if not items:
        return _create_new_preposto(page, nome, cpf, email, telefone)

    _select_preposto(page, nome)
    if not _verify_preposto_selected(page):
        items = _type_and_poll_autocomplete(page, search_term)
        if items:
            _select_preposto(page, nome)
        if not _verify_preposto_selected(page):
            raise Exception("Falha ao selecionar preposto: campo hidden ficou vazio")

    return False


def _type_and_poll_autocomplete(page, search_term) -> list | None:
    page.evaluate("""
        const input = document.querySelector('[id$="pgAutoPreposto_input"]');
        input.value = '';
        input.focus();
        input.click();
    """)
    time.sleep(0.35)
    page.keyboard.type(search_term, delay=65)

    for _ in range(POLL_ATTEMPTS):
        time.sleep(POLL_WAIT)
        try:
            items = page.evaluate("""(() => {
                const panel = document.querySelector('[id$="pgAutoPreposto_panel"]');
                if (!panel) return null;
                const items = Array.from(panel.querySelectorAll('li')).map(i => i.textContent.trim());
                return items.length > 0 ? items : null;
            })()""")
        except Exception:
            return None
        if items:
            return items

    return None


def _select_preposto(page, nome):
    nome_upper = nome.upper()
    page.evaluate(f"""
        const panel = document.querySelector('[id$="pgAutoPreposto_panel"]');
        const items = Array.from(panel.querySelectorAll('li'));
        const target = items.find(i => i.textContent.trim().toUpperCase().includes({repr(nome_upper)}))
                    || items[0];
        target.click();
    """)
    time.sleep(0.6)


def _verify_preposto_selected(page) -> bool:
    val = page.evaluate("""(() => {
        const h = document.querySelector('[id$="pgAutoPreposto_hinput"]');
        return h ? h.value : '';
    })()""")
    return bool(val and val.strip())


# ── Novo cadastro ─────────────────────────────────────────────────────────────

def _create_new_preposto(page, nome, cpf, email, telefone) -> bool:
    result = page.evaluate("""(() => {
        const btn = Array.from(document.querySelectorAll('button, a'))
            .find(b => b.textContent.includes('Novo') && b.textContent.includes('Cadastro'));
        if (btn) { btn.click(); return 'ok'; }
        return 'nao_encontrado';
    })()""")
    if result != "ok":
        raise Exception("Botão 'Novo Cadastro' não encontrado")

    time.sleep(1.5)
    cpf_digits = re.sub(r"\D", "", cpf)

    fill_result = page.evaluate(f"""(() => {{
        const iframe = document.querySelector('iframe[src*="prepostoEdit"]')
                    || document.querySelector('iframe');
        if (!iframe) return 'iframe_nao_encontrado';
        const doc = iframe.contentDocument || iframe.contentWindow.document;

        function fill(el, val) {{
            if (!el) return;
            el.value = val;
            el.dispatchEvent(new Event('input',  {{ bubbles: true }}));
            el.dispatchEvent(new Event('change', {{ bubbles: true }}));
        }}

        const inputs = doc.querySelectorAll(
            'input[type="text"], input:not([type="hidden"]):not([type="submit"]):not([type="button"])'
        );
        fill(inputs[0], {repr(nome)});
        fill(inputs[1], {repr(cpf_digits)});
        fill(inputs[2], {repr(telefone)});
        fill(inputs[3], {repr(email)});
        return 'preenchido';
    }})()""")
    if fill_result == "iframe_nao_encontrado":
        raise Exception("Iframe do formulário Novo Cadastro não encontrado")

    save_result = page.evaluate("""(() => {
        const iframe = document.querySelector('iframe[src*="prepostoEdit"]')
                    || document.querySelector('iframe');
        const doc = iframe.contentDocument || iframe.contentWindow.document;
        const btn = Array.from(doc.querySelectorAll('button, input[type="submit"]'))
            .find(b => b.textContent.includes('Salvar') || b.value === 'Salvar');
        if (btn) { btn.click(); return 'salvo'; }
        return 'nao_encontrado';
    })()""")
    if save_result != "salvo":
        raise Exception("Botão Salvar no formulário de novo preposto não encontrado")

    time.sleep(2)

    items = _type_and_poll_autocomplete(page, " ".join(nome.split()[:3]))
    if items:
        _select_preposto(page, nome)
        if not _verify_preposto_selected(page):
            raise Exception("Preposto cadastrado mas seleção falhou")
    else:
        raise Exception("Preposto foi cadastrado mas não apareceu no autocomplete")

    return True


# ── Relatório Excel ───────────────────────────────────────────────────────────

def _build_report(results: list[dict], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    ws.append(["Número Processo", "Nome Preposto", "CPF", "Email",
               "Status", "Detalhe", "Horário"])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    fills = {
        "OK":            PatternFill("solid", fgColor="C6EFCE"),
        "JÁ CONFIRMADO": PatternFill("solid", fgColor="DDEBF7"),
        "ERRO":          PatternFill("solid", fgColor="FFC7CE"),
    }

    for r in results:
        ws.append([r["numero_processo"], r["nome_preposto"], r["cpf"], r["email"],
                   r["status"], r["detalhe"], r["horario"]])
        for cell in ws[ws.max_row]:
            cell.fill = fills.get(r["status"], fills["ERRO"])

    for col, w in zip("ABCDEFG", [22, 32, 16, 32, 16, 55, 10]):
        ws.column_dimensions[col].width = w

    wb.save(path)
